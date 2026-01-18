"""
Utilidades para exportación de datos de envíos a diferentes formatos
"""
import csv
import io
from datetime import datetime
from typing import List
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


def exportar_envios_excel(envios_queryset, filename='envios_export.xlsx'):
    """
    Exporta un queryset de envíos a formato Excel (.xlsx)
    
    Args:
        envios_queryset: QuerySet de objetos Envio
        filename: Nombre del archivo a generar
        
    Returns:
        HttpResponse con el archivo Excel
    """
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos"
    
    # Definir estilos
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    number_alignment = Alignment(horizontal='right', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Encabezados
    headers = [
        'N° Guía (HAWB)',
        'Destinatario',
        'Cédula',
        'Correo',
        'Teléfono',
        'Ciudad',
        'Estado',
        'Fecha Emisión',
        'Peso Total (kg)',
        'Cantidad Total',
        'Valor Total ($)',
        'Costo Servicio ($)',
        'Observaciones'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    row_num = 2
    for envio in envios_queryset:
        # Obtener información del comprador
        comprador_nombre = envio.comprador.nombre if hasattr(envio, 'comprador') else 'N/A'
        comprador_cedula = envio.comprador.cedula if hasattr(envio, 'comprador') else 'N/A'
        comprador_correo = envio.comprador.correo if hasattr(envio, 'comprador') else 'N/A'
        comprador_telefono = getattr(envio.comprador, 'telefono', 'N/A') if hasattr(envio, 'comprador') else 'N/A'
        comprador_ciudad = getattr(envio.comprador, 'ciudad', 'N/A') if hasattr(envio, 'comprador') else 'N/A'
        
        # Estado legible
        estados_dict = {
            'pendiente': 'Pendiente',
            'en_transito': 'En Tránsito',
            'entregado': 'Entregado',
            'cancelado': 'Cancelado'
        }
        estado_legible = estados_dict.get(envio.estado, envio.estado)
        
        # Fecha formateada
        fecha_emision = envio.fecha_emision.strftime('%d/%m/%Y %H:%M') if envio.fecha_emision else 'N/A'
        
        # Datos de la fila
        row_data = [
            envio.hawb,
            comprador_nombre,
            comprador_cedula,
            comprador_correo,
            comprador_telefono,
            comprador_ciudad,
            estado_legible,
            fecha_emision,
            float(envio.peso_total) if envio.peso_total else 0,
            envio.cantidad_total if envio.cantidad_total else 0,
            float(envio.valor_total) if envio.valor_total else 0,
            float(envio.costo_servicio) if envio.costo_servicio else 0,
            envio.observaciones if envio.observaciones else ''
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Aplicar alineación según el tipo de dato
            if col_num in [9, 10, 11, 12]:  # Columnas numéricas
                cell.alignment = number_alignment
                if col_num in [11, 12]:  # Valores monetarios
                    cell.number_format = '$#,##0.00'
            elif col_num == 7:  # Estado
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
        
        row_num += 1
    
    # Ajustar ancho de columnas
    column_widths = {
        1: 15,  # HAWB
        2: 25,  # Destinatario
        3: 13,  # Cédula
        4: 28,  # Correo
        5: 15,  # Teléfono
        6: 18,  # Ciudad
        7: 15,  # Estado
        8: 18,  # Fecha
        9: 12,  # Peso
        10: 12, # Cantidad
        11: 12, # Valor
        12: 15, # Costo
        13: 30  # Observaciones
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Agregar filtros automáticos
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row_num-1}'
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


def exportar_envios_csv(envios_queryset, filename='envios_export.csv'):
    """
    Exporta un queryset de envíos a formato CSV
    
    Args:
        envios_queryset: QuerySet de objetos Envio
        filename: Nombre del archivo a generar
        
    Returns:
        HttpResponse con el archivo CSV
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Agregar BOM para compatibilidad con Excel en UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
    
    # Encabezados
    writer.writerow([
        'N° Guía (HAWB)',
        'Destinatario',
        'Cédula',
        'Correo',
        'Teléfono',
        'Ciudad',
        'Estado',
        'Fecha Emisión',
        'Peso Total (kg)',
        'Cantidad Total',
        'Valor Total ($)',
        'Costo Servicio ($)',
        'Observaciones'
    ])
    
    # Estados legibles
    estados_dict = {
        'pendiente': 'Pendiente',
        'en_transito': 'En Tránsito',
        'entregado': 'Entregado',
        'cancelado': 'Cancelado'
    }
    
    # Escribir datos
    for envio in envios_queryset:
        comprador_nombre = envio.comprador.nombre if hasattr(envio, 'comprador') else 'N/A'
        comprador_cedula = envio.comprador.cedula if hasattr(envio, 'comprador') else 'N/A'
        comprador_correo = envio.comprador.correo if hasattr(envio, 'comprador') else 'N/A'
        comprador_telefono = getattr(envio.comprador, 'telefono', 'N/A') if hasattr(envio, 'comprador') else 'N/A'
        comprador_ciudad = getattr(envio.comprador, 'ciudad', 'N/A') if hasattr(envio, 'comprador') else 'N/A'
        
        estado_legible = estados_dict.get(envio.estado, envio.estado)
        fecha_emision = envio.fecha_emision.strftime('%d/%m/%Y %H:%M') if envio.fecha_emision else 'N/A'
        
        writer.writerow([
            envio.hawb,
            comprador_nombre,
            comprador_cedula,
            comprador_correo,
            comprador_telefono,
            comprador_ciudad,
            estado_legible,
            fecha_emision,
            f"{envio.peso_total:.2f}" if envio.peso_total else "0.00",
            envio.cantidad_total if envio.cantidad_total else 0,
            f"{envio.valor_total:.2f}" if envio.valor_total else "0.00",
            f"{envio.costo_servicio:.2f}" if envio.costo_servicio else "0.00",
            envio.observaciones if envio.observaciones else ''
        ])
    
    return response


def exportar_envios_pdf(envios_queryset, filename='envios_export.pdf'):
    """
    Exporta un queryset de envíos a formato PDF
    
    Args:
        envios_queryset: QuerySet de objetos Envio
        filename: Nombre del archivo a generar
        
    Returns:
        HttpResponse con el archivo PDF
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Crear el documento PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Título
    title = Paragraph("Reporte de Envíos", title_style)
    elements.append(title)
    
    # Subtítulo con fecha
    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    subtitle = Paragraph(f"Generado el {fecha_actual} | Total de envíos: {envios_queryset.count()}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 12))
    
    # Estados legibles
    estados_dict = {
        'pendiente': 'Pendiente',
        'en_transito': 'En Tránsito',
        'entregado': 'Entregado',
        'cancelado': 'Cancelado'
    }
    
    # Preparar datos para la tabla
    data = [[
        'HAWB',
        'Destinatario',
        'Ciudad',
        'Estado',
        'Fecha',
        'Peso\n(kg)',
        'Valor\n($)',
        'Costo\n($)'
    ]]
    
    for envio in envios_queryset:
        comprador_nombre = envio.comprador.nombre[:20] if hasattr(envio, 'comprador') else 'N/A'
        comprador_ciudad = getattr(envio.comprador, 'ciudad', 'N/A')[:15] if hasattr(envio, 'comprador') else 'N/A'
        estado_legible = estados_dict.get(envio.estado, envio.estado)
        fecha_emision = envio.fecha_emision.strftime('%d/%m/%Y') if envio.fecha_emision else 'N/A'
        
        data.append([
            envio.hawb[:12],
            comprador_nombre,
            comprador_ciudad,
            estado_legible,
            fecha_emision,
            f"{envio.peso_total:.2f}" if envio.peso_total else "0.00",
            f"{envio.valor_total:.2f}" if envio.valor_total else "0.00",
            f"{envio.costo_servicio:.2f}" if envio.costo_servicio else "0.00"
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[
        1.1*inch,  # HAWB
        1.5*inch,  # Destinatario
        1.0*inch,  # Ciudad
        0.9*inch,  # Estado
        0.8*inch,  # Fecha
        0.6*inch,  # Peso
        0.6*inch,  # Valor
        0.6*inch   # Costo
    ])
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Cuerpo
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (4, -1), 'LEFT'),
        ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        
        # Filas alternas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))
    
    elements.append(table)
    
    # Pie de página con totales
    elements.append(Spacer(1, 20))
    
    # Calcular totales
    total_peso = sum([float(e.peso_total or 0) for e in envios_queryset])
    total_valor = sum([float(e.valor_total or 0) for e in envios_queryset])
    total_costo = sum([float(e.costo_servicio or 0) for e in envios_queryset])
    
    totales_text = f"""
    <b>RESUMEN GENERAL:</b><br/>
    Peso Total: {total_peso:.2f} kg | 
    Valor Total: ${total_valor:.2f} | 
    Costo Total Servicio: ${total_costo:.2f}
    """
    
    totales_style = ParagraphStyle(
        'Totales',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#2c3e50'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    totales_para = Paragraph(totales_text, totales_style)
    elements.append(totales_para)
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener el valor del buffer
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


def generar_comprobante_envio(envio, filename='comprobante_envio.pdf'):
    """
    Genera un comprobante detallado de un envío específico en formato PDF
    
    Args:
        envio: Objeto Envio
        filename: Nombre del archivo a generar
        
    Returns:
        HttpResponse con el comprobante PDF
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título principal
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("COMPROBANTE DE ENVÍO", title_style))
    elements.append(Spacer(1, 20))
    
    # Información del envío
    info_style = styles['Normal']
    info_style.fontSize = 11
    
    # Número de guía destacado
    hawb_style = ParagraphStyle(
        'HAWB',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#3498db'),
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph(f"N° GUÍA: {envio.hawb}", hawb_style))
    elements.append(Spacer(1, 15))
    
    # Información del destinatario
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("INFORMACIÓN DEL DESTINATARIO", section_style))
    
    if hasattr(envio, 'comprador') and envio.comprador:
        # Crear estilo para texto que se ajusta automáticamente
        nombre_style = ParagraphStyle(
            'NombreStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            leading=12,
            wordWrap='CJK'  # Permite ajuste de texto
        )
        
        # Usar Paragraph para que el texto se ajuste automáticamente
        nombre_paragraph = Paragraph(envio.comprador.nombre or 'N/A', nombre_style)
        cedula_paragraph = Paragraph(envio.comprador.cedula or 'N/A', nombre_style)
        correo_paragraph = Paragraph(envio.comprador.correo or 'N/A', nombre_style)
        telefono_paragraph = Paragraph(getattr(envio.comprador, 'telefono', 'N/A') or 'N/A', nombre_style)
        ciudad_paragraph = Paragraph(getattr(envio.comprador, 'ciudad', 'N/A') or 'N/A', nombre_style)
        
        destinatario_data = [
            ['Nombre:', nombre_paragraph],
            ['Cédula:', cedula_paragraph],
            ['Correo:', correo_paragraph],
            ['Teléfono:', telefono_paragraph],
            ['Ciudad:', ciudad_paragraph]
        ]
        
        dest_table = Table(destinatario_data, colWidths=[2*inch, 4*inch])
        dest_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#7f8c8d')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alinear al inicio verticalmente
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)  # Opcional: agregar bordes sutiles
        ]))
        elements.append(dest_table)
    
    elements.append(Spacer(1, 20))
    
    # Información del envío
    elements.append(Paragraph("DETALLES DEL ENVÍO", section_style))
    
    estados_dict = {
        'pendiente': 'Pendiente',
        'en_transito': 'En Tránsito',
        'entregado': 'Entregado',
        'cancelado': 'Cancelado'
    }
    
    estado_legible = estados_dict.get(envio.estado, envio.estado)
    fecha_emision = envio.fecha_emision.strftime('%d/%m/%Y %H:%M') if envio.fecha_emision else 'N/A'
    
    envio_data = [
        ['Estado:', estado_legible],
        ['Fecha de Emisión:', fecha_emision],
        ['Peso Total:', f"{envio.peso_total:.2f} kg" if envio.peso_total else "0.00 kg"],
        ['Cantidad Total:', str(envio.cantidad_total if envio.cantidad_total else 0)],
        ['Valor Total:', f"${envio.valor_total:.2f}" if envio.valor_total else "$0.00"],
        ['Costo del Servicio:', f"${envio.costo_servicio:.2f}" if envio.costo_servicio else "$0.00"]
    ]
    
    envio_table = Table(envio_data, colWidths=[2*inch, 4*inch])
    envio_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#7f8c8d')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT')
    ]))
    elements.append(envio_table)
    
    elements.append(Spacer(1, 20))
    
    # Productos (si existen)
    if hasattr(envio, 'productos') and envio.productos.exists():
        elements.append(Paragraph("PRODUCTOS", section_style))
        
        productos_data = [['Descripción', 'Categoría', 'Peso (kg)', 'Cantidad', 'Valor ($)']]
        
        for producto in envio.productos.all():
            productos_data.append([
                producto.descripcion[:30],
                producto.get_categoria_display() if hasattr(producto, 'get_categoria_display') else producto.categoria,
                f"{producto.peso:.2f}" if producto.peso else "0.00",
                str(producto.cantidad),
                f"{producto.valor:.2f}" if producto.valor else "0.00"
            ])
        
        prod_table = Table(productos_data, colWidths=[2.2*inch, 1.3*inch, 0.9*inch, 0.9*inch, 0.9*inch])
        prod_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Cuerpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        elements.append(prod_table)
    
    # Observaciones
    if envio.observaciones:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("OBSERVACIONES", section_style))
        obs_text = Paragraph(envio.observaciones, styles['Normal'])
        elements.append(obs_text)
    
    # Pie de página
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer_text = f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    elements.append(Paragraph(footer_text, footer_style))
    
    # Construir PDF
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response























