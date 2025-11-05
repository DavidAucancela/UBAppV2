"""
Comando de Django para generar plantilla de importación de Excel
"""
from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class Command(BaseCommand):
    help = 'Genera una plantilla de ejemplo para importación de envíos desde Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--with-data',
            action='store_true',
            help='Incluir datos de ejemplo en la plantilla',
        )
        parser.add_argument(
            '--output',
            type=str,
            default='plantilla_importacion_envios.xlsx',
            help='Nombre del archivo de salida',
        )

    def handle(self, *args, **options):
        with_data = options['with_data']
        output_file = options['output']

        self.stdout.write(self.style.SUCCESS('Generando plantilla de importación...'))

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos de Envíos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = [
            'HAWB',
            'Peso Total',
            'Cantidad Total',
            'Valor Total',
            'Estado',
            'Observaciones',
            'Descripción Producto',
            'Peso Producto',
            'Cantidad Producto',
            'Valor Producto',
            'Categoría'
        ]

        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ajustar ancho de columnas
        column_widths = [15, 12, 15, 12, 15, 30, 30, 15, 18, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Agregar datos de ejemplo si se solicita
        if with_data:
            datos_ejemplo = [
                ['HAWB001', 5.5, 2, 150.00, 'pendiente', 'Envío urgente', 'Laptop Dell Inspiron', 2.5, 1, 100.00, 'electronica'],
                ['HAWB002', 1.2, 3, 45.50, 'pendiente', '', 'Camiseta Nike', 0.4, 3, 45.50, 'ropa'],
                ['HAWB003', 3.0, 1, 80.00, 'en_transito', 'Frágil', 'Cafetera Oster', 3.0, 1, 80.00, 'hogar'],
                ['HAWB004', 2.8, 5, 125.00, 'pendiente', '', 'Balón de fútbol Adidas', 0.56, 5, 125.00, 'deportes'],
                ['HAWB005', 0.8, 2, 35.00, 'entregado', 'Entregado el 15/10', 'Audífonos Bluetooth', 0.4, 2, 35.00, 'electronica'],
            ]

            for row_num, datos in enumerate(datos_ejemplo, 2):
                for col_num, valor in enumerate(datos, 1):
                    ws.cell(row=row_num, column=col_num).value = valor

        # Crear hoja de instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        
        instrucciones = [
            ["GUÍA DE USO DE LA PLANTILLA DE IMPORTACIÓN", ""],
            ["", ""],
            ["Columnas Disponibles:", ""],
            ["", ""],
            ["HAWB", "OBLIGATORIO - Número único de guía de envío"],
            ["Peso Total", "Peso total del envío en kilogramos"],
            ["Cantidad Total", "Cantidad total de productos en el envío"],
            ["Valor Total", "Valor total del envío en dólares (USD)"],
            ["Estado", "Estado del envío: pendiente, en_transito, entregado, cancelado"],
            ["Observaciones", "Notas o comentarios adicionales"],
            ["Descripción Producto", "Descripción detallada del producto"],
            ["Peso Producto", "Peso individual del producto en kilogramos"],
            ["Cantidad Producto", "Cantidad del producto"],
            ["Valor Producto", "Valor del producto en dólares (USD)"],
            ["Categoría", "Categoría: electronica, ropa, hogar, deportes, otros"],
            ["", ""],
            ["NOTAS IMPORTANTES:", ""],
            ["", ""],
            ["1. El campo HAWB es OBLIGATORIO y debe ser único", ""],
            ["2. Los valores numéricos deben usar punto (.) como separador decimal", ""],
            ["3. Las categorías válidas son: electronica, ropa, hogar, deportes, otros", ""],
            ["4. Los estados válidos son: pendiente, en_transito, entregado, cancelado", ""],
            ["5. Si no especifica estado, se usará 'pendiente' por defecto", ""],
            ["6. Si no especifica categoría, se usará 'otros' por defecto", ""],
            ["", ""],
            ["PROCESO DE IMPORTACIÓN:", ""],
            ["", ""],
            ["1. Complete la hoja 'Datos de Envíos' con su información", ""],
            ["2. Guarde el archivo en formato Excel (.xlsx)", ""],
            ["3. Suba el archivo en el módulo de importación", ""],
            ["4. Mapee las columnas si es necesario", ""],
            ["5. Revise la validación de datos", ""],
            ["6. Seleccione los registros a importar", ""],
            ["7. Complete la importación", ""],
            ["", ""],
            [f"Plantilla generada: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", ""],
        ]

        for row_num, (col1, col2) in enumerate(instrucciones, 1):
            ws_instrucciones.cell(row=row_num, column=1).value = col1
            ws_instrucciones.cell(row=row_num, column=2).value = col2

        # Formato especial para el título
        ws_instrucciones.cell(row=1, column=1).font = Font(bold=True, size=14, color="4472C4")
        
        # Formato para secciones
        for row in [3, 17, 26]:
            ws_instrucciones.cell(row=row, column=1).font = Font(bold=True, size=12)

        # Ajustar anchos
        ws_instrucciones.column_dimensions['A'].width = 35
        ws_instrucciones.column_dimensions['B'].width = 60

        # Guardar archivo
        try:
            wb.save(output_file)
            self.stdout.write(
                self.style.SUCCESS(
                    f'✅ Plantilla generada exitosamente: {output_file}'
                )
            )
            
            if with_data:
                self.stdout.write(
                    self.style.SUCCESS(
                        f'   Incluye {len(datos_ejemplo)} registros de ejemplo'
                    )
                )
            
            self.stdout.write('')
            self.stdout.write('Puede usar esta plantilla para importar envíos al sistema.')
            self.stdout.write('Consulte la hoja "Instrucciones" para más información.')
            
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'❌ Error al guardar el archivo: {str(e)}')
            )
            return

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('¡Proceso completado!'))






